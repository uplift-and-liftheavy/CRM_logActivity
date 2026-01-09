import time
import os
import getpass
import logging
import ctypes
import openpyxl as xl

# Selenium Imports
from selenium import webdriver
# from selenium.webdriver.chrome.service import Service as ChromeService  # not needed with Selenium Manager
# from webdriver_manager.chrome import ChromeDriverManager               # removed; Selenium Manager will handle drivers
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import (
    NoSuchElementException,
    StaleElementReferenceException,
    ElementClickInterceptedException,
    TimeoutException
)
from selenium.webdriver.common.keys import Keys

###############################################################################
# GLOBALS
###############################################################################

# Basic wait durations
SHORT_WAIT = 0.5
MED_WAIT = 1
LONG_WAIT = 3

# Prevent system from going to sleep (Windows only)
ctypes.windll.kernel32.SetThreadExecutionState(0x80000002)

# Setup logging
logging.basicConfig(
    filename='error.log',
    level=logging.DEBUG,
    format='%(asctime)s %(levelname)s %(name)s %(message)s'
)
logger = logging.getLogger(__name__)

###############################################################################
# WEB DRIVER SETUP
###############################################################################
# Let Selenium Manager locate/download the appropriate ChromeDriver automatically.
driver = webdriver.Chrome()
driver.implicitly_wait(10)
driver.set_script_timeout(10)
driver.set_page_load_timeout(10)
action = ActionChains(driver)

###############################################################################
# EXCEL SETUP (Example)
###############################################################################

pathUsername = getpass.getuser()
reportPath = os.path.join(
    "C:",
    os.path.sep,
    "Users",
    pathUsername,
    "Documents",
    "Wago Tools",
    "SAP - Log Activity",
    "Log Activity.xlsx"
)

report = xl.load_workbook(reportPath)
sheet1 = report["Activity Report"]

columnAccountID = 2
columnContactName = 3
columnContactID = 4
columnDescription = 5
columnStartDateTime = 6
columnNotes = 7
columnLoggedSuccessfully = 8

###############################################################################
# HELPER FUNCTIONS
###############################################################################

def check_for_element(selector, description, maxAttempts=3):
    """
    Tries to locate an element up to 'maxAttempts' times.
    Shows a short console message on failure, logs full stacktrace to log file.
    """
    attempts = 0
    while attempts < maxAttempts:
        try:
            WebDriverWait(driver, 1).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, selector))
            )
            print(f"Located '{description}' successfully on attempt {attempts+1}.")
            return True
        except (TimeoutException, NoSuchElementException) as e:
            attempts += 1
            short_exc_name = type(e).__name__
            print(
                f"Attempt {attempts} failed to find '{description}' "
                f"(Exception: {short_exc_name})"
            )
            logger.exception(
                f"[checkForElement] Full exception details for '{description}' "
                f"on attempt {attempts}:"
            )
            time.sleep(SHORT_WAIT)
            if attempts == maxAttempts:
                print(
                    f"Max attempts reached, could not locate the element '{description}'."
                )
                return False


def wait_for_visibility(selector, wait_time=5):
    """Wait for the element to be visible."""
    WebDriverWait(driver, wait_time).until(
        EC.visibility_of(driver.find_element(By.CSS_SELECTOR, selector))
    )


def wait_for_clickable(selector, wait_time=5):
    """Wait for the element to be clickable."""
    WebDriverWait(driver, wait_time).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
    )


def click_element(
    currentElementSelector,
    elementDescription,
    maxAttempts=10,
    stopAfterClickBeforeCheckClick=SHORT_WAIT,
    stopAfterClickedSuccessfully=SHORT_WAIT
):
    """
    Attempts to load and click an element up to 'maxAttempts' times.
    Shows minimal console output, logs full exceptions to the log file.
    """
    print(f"Attempting to click '{elementDescription}' (up to {maxAttempts} attempts).")

    attempts = 0
    while attempts < maxAttempts:
        attempts += 1
        try:
            wait_for_clickable(currentElementSelector, 5)
            element = driver.find_element(By.CSS_SELECTOR, currentElementSelector)
            element.click()
            print(f"The element '{elementDescription}' was clicked successfully.")
            time.sleep(stopAfterClickBeforeCheckClick)
            time.sleep(stopAfterClickedSuccessfully)
            return True  # success
        except (TimeoutException, NoSuchElementException, ElementClickInterceptedException) as e:
            short_exc_name = type(e).__name__
            print(
                f"Attempt {attempts} failed to click '{elementDescription}' "
                f"(Exception: {short_exc_name})"
            )
            logger.exception(
                f"[click_element] Full exception details for '{elementDescription}' "
                f"on attempt {attempts}:"
            )
        except StaleElementReferenceException as e:
            short_exc_name = type(e).__name__
            print(
                f"Stale reference for '{elementDescription}' "
                f"on attempt {attempts} (Exception: {short_exc_name})"
            )
            logger.exception(
                f"[click_element] Full exception details for '{elementDescription}' "
                f"on attempt {attempts}:"
            )
            # JS fallback
            try:
                element = driver.find_element(By.CSS_SELECTOR, currentElementSelector)
                driver.execute_script("arguments[0].click();", element)
                print(f"'{elementDescription}' clicked with JS fallback.")
                time.sleep(stopAfterClickBeforeCheckClick)
                time.sleep(stopAfterClickedSuccessfully)
                return True
            except Exception as e2:
                logger.error(
                    f"JS fallback also failed for '{elementDescription}': {e2}"
                )

        # If we get here, click wasn't successful this attempt
        if attempts == maxAttempts:
            print(f"Max attempts reached. Could not click '{elementDescription}'.\n")
            return False

    # If we exit while, all attempts used up
    time.sleep(stopAfterClickedSuccessfully)
    return False


def openAccountPage():
    """Try #__div6, else expand 'Customers' first if not found."""
    print("Opening Account Page from home screen...\n")
    sel = '#__div6'
    try:
        wait_for_visibility(sel)
        wait_for_clickable(sel)
        click_element(sel, "Account (div6)")
    except (NoSuchElementException, TimeoutException):
        print("Couldn't find '#__div6' directly. Expanding 'Customers' first...\n")
        sel_cust = '#__panel1 [title="Expand/Collapse"]'
        click_element(sel_cust, "Customers Expand/Collapse", 5)
        time.sleep(SHORT_WAIT)
        click_element(sel, "Account (div6) again", 5)
    time.sleep(SHORT_WAIT)

def openContactPage():
    """Try #__div6, else expand 'Customers' first if not found."""
    print("Opening Account Page from home screen...\n")
    sel = 'a[title="Contacts"]'
    #'div.sapMPanelContent.sapMPanelBGTranslucent.sapMPanelExpandablePart > div:nth-of-type(3)'
    #'#__div6'
    try:
        wait_for_visibility(sel)
        wait_for_clickable(sel)
        click_element(sel, "Contacts")
    except (NoSuchElementException, TimeoutException):
        print("Couldn't find '#__div6' directly. Expanding 'Customers' first...\n")
        sel_cust = '#__panel1 [title="Expand/Collapse"]'
        click_element(sel_cust, "Customers Expand/Collapse", 5)
        time.sleep(SHORT_WAIT)
        click_element(sel, "Account (div6) again", 5)
    time.sleep(LONG_WAIT)    

def set_contacts_to_all_WAGO():
    print('Set Contacts to "All Contacts" Not just Mine')
    # Click Dropdown
    currentElementSelector_typeDropdown = (
        'div.sapClientMCoreBaseVariantManagement.sapClientMCoreBaseVariantManagementMulti.sapUiCompVarMngmt.sapUiVariantManagementNoToolbar.sapMBarChild[title="Select View"]'
        ' div.sapUiHLayout.sapUiHLayoutNoWrap.sapUICompVarMngmtLayout'
        ' div.sapUiHLayoutChildWrapper:nth-of-type(3)'
        ' button.sapMBtnBase.sapMBtn.sapUICompVarMngmtTriggerBtn.sapMTitleStyleH4[title="Select View"]'
        ' span.sapMBtnInner.sapMBtnHoverable.sapMFocusable.sapMBtnIconFirst.sapMBtnTransparent'
        ' span.sapUiIcon.sapUiIconMirrorInRTL.sapMBtnCustomIcon.sapMBtnIcon.sapMBtnIconLeft'
        )
    elementDescription_typeDropdown = 'Select View (Mine or All WAGO)'
    click_element(
        currentElementSelector_typeDropdown, 
        elementDescription_typeDropdown, 
        maxAttempts=5,
        stopAfterClickBeforeCheckClick=SHORT_WAIT,
        stopAfterClickedSuccessfully=SHORT_WAIT
    )
    
    currentElementSelector = (
        'div.sapMPopover.sapMPopoverWithBar.sapMPopoverWithoutSubHeader.sapMPopoverPage.sapMPopoverWithoutFooter.sapMPopoverVerScrollDisabled.sapMPopoverHorScrollDisabled.sapMPopup-CTX.sapMResponsivePopover.sapUICompVarMngmtPopover.sapMPopoverWithHeaderCont.sapMResponsivePopover.sapUICompVarMngmtPopover.sapMPopoverWithHeaderCont.sapUiShd'
        ' div.sapMPopoverCont'
        ' section.sapMPageEnableScrolling'
        ' ul.sapMSelectList'
        ' li:nth-of-type(3)'
        )
    elementDescription = 'All WAGO Contacts)'
    click_element(
        currentElementSelector, 
        elementDescription, 
        maxAttempts=5,
        stopAfterClickBeforeCheckClick=SHORT_WAIT,
        stopAfterClickedSuccessfully=SHORT_WAIT
    )

def click_search_and_reset_field():
    """Reset the search field if necessary."""
    print("Resetting the search field (if available).")
    try:
        currentElementSelector = (
            '.sapMIBar.sapMTB.sapMTBNewFlex.sapMTBInactive.sapMTBStandard.sapMTB-Auto-CTX.sapClientMALPToolbarRow.sapMBarChild.sapMTBShrinkItem'
            ' > button[title="Search"]'
        )
        elementDescription = 'Search Reset'
        if check_for_element(currentElementSelector, elementDescription, maxAttempts=2):
            click_element(
                currentElementSelector, 
                elementDescription, 
                maxAttempts=2,
                stopAfterClickBeforeCheckClick=SHORT_WAIT,
                stopAfterClickedSuccessfully=SHORT_WAIT
            )
        else: 
            try:
                currentElementSelector = '[data-sap-ui-icon-content=""]'
                elementDescription = 'Magnifying Glass'
                click_element(
                    currentElementSelector, 
                    elementDescription, 
                    maxAttempts=2,
                    stopAfterClickBeforeCheckClick=SHORT_WAIT,
                    stopAfterClickedSuccessfully=SHORT_WAIT
                )
                
                currentElementSelector = '[data-help-id="m8HSa5dpNqg1z8nY0idosG-searchField"] [title="Reset"]'
                elementDescription = 'Search Reset'
                click_element(
                    currentElementSelector, 
                    elementDescription, 
                    maxAttempts=2,
                    stopAfterClickBeforeCheckClick=SHORT_WAIT,
                    stopAfterClickedSuccessfully=SHORT_WAIT
                )
            except:
                pass
    except:
        print("Couldn't reset search (this is okay). Continuing on...\n")

def enterAccountNumberAndSearch(account_number):
    """Click the mag glass, then type the account number."""
    print(f"Entering account number: {account_number}")
    mag_glass_selector = '[data-sap-ui-icon-content=""]'
    click_element(mag_glass_selector, "Magnifying Glass", 2)
    search_field_selector = '[placeholder="Search"]'
    wait_for_visibility(search_field_selector)
    wait_for_clickable(search_field_selector)
    sf = driver.find_element(By.CSS_SELECTOR, search_field_selector)
    sf.clear()
    sf.send_keys(account_number)
    time.sleep(SHORT_WAIT)

def verifyAccountNumber(expected_account_number, retry_count=0, max_retries=5):
    """
    Checks if the displayed account number matches the expected.
    If mismatch, resets the search and retries (up to max_retries).
    """
    if retry_count >= max_retries:
        print("Max retries reached. Verification failed.\n")
        return None

    account_display_selector = (
        '.sapMScrollContH.sapMScrollCont.sapClientMALPScrollContainer'
        ' .sapMList.sapMListBGTranslucent.sapMListTblCnt.sapClientMColFreeze.'
        'sapClientMTable.display-style-table.saplsui-widthauto.selectionModeDisabled'
        ' table:nth-of-type(2)'
        ' tbody:nth-of-type(1)'
        ' tr:nth-of-type(1)'
        ' td:nth-of-type(3)'
        ' div:nth-of-type(1)'
        ' span:nth-of-type(1)'
    )
    try:
        time.sleep(MED_WAIT)
        elem = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, account_display_selector))
        )
        displayed_account_number = elem.text.strip()
        expected_account_number = str(expected_account_number).strip()
        if displayed_account_number == expected_account_number:
            print(f"Account number verified successfully: {displayed_account_number}\n")
            return displayed_account_number
        else:
            print("Account number mismatch, retrying...\n")
            click_search_and_reset_field()
            time.sleep(SHORT_WAIT)
            enterAccountNumberAndSearch(expected_account_number)
            return verifyAccountNumber(expected_account_number, retry_count+1)
    except NoSuchElementException:
        print("Account number display not found, resetting & retrying...\n")
        click_search_and_reset_field()
        time.sleep(SHORT_WAIT)
        enterAccountNumberAndSearch(expected_account_number)
        time.sleep(MED_WAIT)
        return verifyAccountNumber(expected_account_number, retry_count+1)
    except TimeoutException:
        print("Timed out waiting for the account number display element.\n")
        return None

def verifyContactID(expected_contact_id, retry_count=0, max_retries=5):
    """
    Checks if the displayed account number matches the expected.
    If mismatch, resets the search and retries (up to max_retries).
    """

    contact_id_selector = (
        '.fully-loaded.sapClientMALP-table-body.sapClientIsTableView.sapClientTableSlimScroll.sapMListUl.sapMListTbl.sapMListShowSeparatorsAll.sapMListModeMultiSelect.sapClientTableSlimScroll-hide'
        ' tbody'
        ' > tr > td:nth-of-type(3) > div > span'
    )
    while retry_count < max_retries:
        try:
            time.sleep(MED_WAIT)
            elem = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, contact_id_selector))
            )
            displayed_contact_id = elem.text.strip()
            expected_contact_id = str(expected_contact_id).strip()
            if displayed_contact_id == expected_contact_id:
                print(f"Contact ID verified successfully: {displayed_contact_id}\n")
                return displayed_contact_id
            else:
                print("Account number mismatch, retrying...\n")
        except (NoSuchElementException, TimeoutException):
            retry_count+=1
            
    if retry_count == max_retries:
        print("Max retries reached. Verification failed.\n")
        return None

def verify_name_flexible(input_name, found_name):
    """Compare two names ignoring case, commas, splitting by whitespace."""
    input_name_set = set(input_name.lower().replace(",", "").split())
    found_name_set = set(found_name.lower().replace(",", "").split())
    if input_name_set == found_name_set:
        print(f"Names match. Input='{input_name}' Found='{found_name}'.\n")
        return True
    else:
        print(f"Names do not match. Input='{input_name}' Found='{found_name}'.\n")
        return False



###############################################################################
# EXAMPLE MAIN LOOP
###############################################################################

def process_one_row(row_index):
    """
    Replicates the same flow as your original code for a single row:
      - openAccountPage
      - resetSearch
      - enterAccountNumberAndSearch
      - verifyAccountNumber
      - open account & popup
      - go to Visits
      - create New Visit
      - fill subject, contact, times, notes
      - save & close
    """
    print(f"\n--- Processing row {row_index} ---\n")

    # 1) Open Account Page
    openContactPage()
    set_contacts_to_all_WAGO()

    # 2) Read the Account ID from Excel
    accountCell = sheet1.cell(row_index, columnAccountID)
    if accountCell.value is None:
        print("No AccountID found in this row, skipping...\n")
        return
    account_id = str(accountCell.value).strip()
    
    # 2b) Read the Contact ID from Excel 
    contactCell = sheet1.cell(row_index, columnContactID)
    if contactCell.value is None:
        print("No AccountID found in this row, skipping...\n")
        return
    contact_id = str(contactCell.value).strip()
    # 3) Reset & search
    click_search_and_reset_field()
    
    time.sleep(SHORT_WAIT)
    enterAccountNumberAndSearch(contact_id)
    time.sleep(MED_WAIT)
    expected_contact_id = contact_id
    verified_contact_id = verifyContactID(expected_contact_id, retry_count=0, max_retries=3)
    #verified_account = verifyAccountNumber(contact_id)
    '''
    if not verified_account:
        sheet1.cell(row=row_index, column=columnLoggedSuccessfully).value = "ERROR: Account not found."
        report.save(reportPath)
        print(f"Account '{account_id}' was not verified. Moving on.\n")
        return
    '''
    if not verified_contact_id:
        sheet1.cell(row=row_index, column=columnLoggedSuccessfully).value = "ERROR: Contact not found."
        report.save(reportPath)
        print(f"Contact '{contact_id}' was not verified. Moving on.\n")
        return

    # 4) If verified, open the actual account link
    open_selector = (
        '.sapMLIB.sapMLIB-CTX.sapMLIBShowSeparator.sapMLIBTypeInactive.'
        'sapMLIBActionable.sapMLIBHoverable.sapMLIBFocusable.sapMListTblRow '
        'td:nth-of-type(4) > div > a'
    )
    click_element(open_selector, "Open Account Link", 10)
    time.sleep(MED_WAIT)

    # 4a) "Open Object Detail" popup
    popup_selector = '.sapClientMQVTIOPeningButtonClass [title="Open Object Detail"]'
    click_element(popup_selector, "Open Account Popup", 10)
    
    time.sleep(LONG_WAIT)

    # 4b) Open Account from Contact
    currentElementSelector = (
        'div.sapMList.sapMListBGTranslucent.sapMListTblCnt.sapClientMColFreeze.sapClientMTable.display-style-table.saplsui-widthauto.selectionModeDisabled.editMode'
        ' tr.sapMLIB.sapMLIB-CTX.sapMLIBShowSeparator.sapMLIBTypeInactive.sapMLIBActionable.sapMLIBHoverable.sapMLIBSelected.sapMLIBFocusable.sapMListTblRow'
        ' td:nth-of-type(3)'
        ' > div > a'
        )
    elementDescription = 'Account'
    click_element(currentElementSelector, elementDescription, 10)
    
    # 4a) "Open Object Detail" popup
    popup_selector = '.sapClientMQVTIOPeningButtonClass d[title="Open Object Detail"]'
    currentElementSelector = (
        'div.sapMPopover.sapMPopoverWithoutBar.sapMPopoverWithoutSubHeader.sapMPopoverWithoutFooter.sapMPopup-CTX.sapMQuickView.sapClientMQVPopover.sapClientMQVOverlay.sapClientMQT.sapClientMQVOverlayMultiTabEnabled.sapMQuickViewHideFooter.sapMQuickView.sapClientMQVPopover.sapClientMQVOverlay.sapClientMQT.sapClientMQVOverlayMultiTabEnabled.sapMQuickViewHideFooter.sapUiShd'
        ' div.sapClientBaseControlsSimpleVLayout.sapClientMQVHeaderWrapper.sapClientMQTHeader'
        ' div.sapClientMQVHRightIcons'
        ' div.sapClientMQVTIOPeningButtonClass'
        ' span[title="Open Object Detail"]'
        )
    elementDescription = 'Open Object Detail (popup)'
    click_element(currentElementSelector, elementDescription, 10)
    time.sleep(LONG_WAIT)
    
    # 5) Click "Visits" Under Account
    currentElementSelector = (
        '.sapMITB.sapMITBNoContentPadding.sapMITBBackgroundDesignSolid.'
        'saplsui-widthauto.sapClientMODNav .sapMITBScrollContainer > div '
        '> div:nth-of-type(8) > div:nth-of-type(1)'
    )
    elementDescription = 'Visits'
    click_element(
        currentElementSelector, elementDescription,
        maxAttempts=10,
        stopAfterClickBeforeCheckClick=SHORT_WAIT,
        stopAfterClickedSuccessfully=SHORT_WAIT
    )
    
    # 6) New Visit
    new_visit_sel = (
        '.sapUiVltCell.sapuiVltCell.sapClientMALPToolbarParent '
        '.sapMBtnBase.sapMBtn.width-button-form.sapMBarChild '
        '.sapMBtnContent'
    )
    click_element(new_visit_sel, "New Visit", 5)
    time.sleep(LONG_WAIT)

    # 7) Dropdown -> "Customer Visit"
    currentElementSelector = (
        '.sapClientBaseControlsCoreOberonComposite.sapClientMComboBox.'
        'width-code-form-value-codeandvalue .sapMInputBaseIconContainer '
        '[data-sap-ui-icon-content=""]'
    )
    elementDescription = 'Dropdown for VisitType'
    click_element(
        currentElementSelector, 
        elementDescription,
        maxAttempts=10,
        stopAfterClickBeforeCheckClick=SHORT_WAIT,
        stopAfterClickedSuccessfully=SHORT_WAIT
    )
    
    currentElementSelector = (
        '.sapMListItems.sapMListUl.sapMListShowSeparatorsNone.'
        'sapMListModeSingleSelectMaster li:nth-of-type(1)'
    )
    elementDescription = 'Customer Visit'
    click_element(
        currentElementSelector, 
        elementDescription,
        maxAttempts=10,
        stopAfterClickBeforeCheckClick=SHORT_WAIT,
        stopAfterClickedSuccessfully=SHORT_WAIT
    )    

    # 8) Subject
    currentElementSelector = (
        '.sapClientBaseControlsCoreOberonComposite.width-name-form-value '
        '.sapMInputBaseNoWidth.sapMInputBase.sapMInputBaseHeightMargin.sapMInput '
        '.sapMInputBaseContentWrapper .sapMInputBaseInner'
    )
    elementDescription = 'Subject'
    click_element(
        currentElementSelector, 
        elementDescription,
        maxAttempts=10,
        stopAfterClickBeforeCheckClick=SHORT_WAIT,
        stopAfterClickedSuccessfully=SHORT_WAIT
    )
    
    subject_cell = sheet1.cell(row_index, columnDescription)
    if subject_cell.value:
        subject_text = subject_cell.value
        el = driver.find_element(By.CSS_SELECTOR, currentElementSelector)
        el.clear()
        time.sleep(SHORT_WAIT)
        action.move_to_element(el).click().send_keys_to_element(el, subject_text).perform()
    
    time.sleep(MED_WAIT)
    # 9) Contact Name
    currentElementSelector = (
        '.sapUiFormResGrid.sapUiFormBackgrTranslucent > div > div:nth-of-type(5) '
        '.sapMInputBaseContentWrapper .sapMInputBaseIconContainer'
    )
    elementDescription = 'Contact Name'
    click_element(
        currentElementSelector, 
        elementDescription,
        maxAttempts=10,
        stopAfterClickBeforeCheckClick=SHORT_WAIT,
        stopAfterClickedSuccessfully=MED_WAIT
    )

    # 9a) Magnifying Glass
    currentElementSelector = (
        '.sapMDialogSection .sapUiVltCell.sapuiVltCell.sapClientMALPToolbarParent '
        '[title="Search"] > span:nth-of-type(1)'
    )
    elementDescription = "mag glass"
    click_element(
        currentElementSelector, 
        elementDescription,
        maxAttempts=5,
        stopAfterClickBeforeCheckClick=SHORT_WAIT,
        stopAfterClickedSuccessfully=MED_WAIT
    )
    time.sleep(MED_WAIT)
    contact_text = sheet1.cell(row_index, columnContactName).value
    time.sleep(MED_WAIT)
    if contact_text:
        action.reset_actions()
        action.send_keys(contact_text).perform()
        time.sleep(MED_WAIT)

        # 9b) Check search result
        currentElementSelector = (
            '.sapClientBaseControlsCoreOberonComposite.width-name-cellRenderer-value.'
            'sapClientMColumnSize-name '
            '.sapMText.sapUiSelectable.sapMTextNoWrap.sapMTextMaxWidth.saplsui-maxwidth100'
        )
        elementDescription = 'Search Result (contact)'
        attempts_verify = 0
        maxVerif = 7
        while attempts_verify < maxVerif:
            attempts_verify += 1
            wait_for_visibility(currentElementSelector)
            wait_for_clickable(currentElementSelector)
            found_element = driver.find_element(By.CSS_SELECTOR, currentElementSelector)
            found_name = found_element.text
            time.sleep(MED_WAIT)
            if verify_name_flexible(contact_text, found_name):                
                click_element(
                    currentElementSelector, 
                    elementDescription,
                    maxAttempts=5,
                    stopAfterClickBeforeCheckClick=SHORT_WAIT,
                    stopAfterClickedSuccessfully=SHORT_WAIT
                )
                break
            else:
                time.sleep(SHORT_WAIT)
                pass

    # 10) Start Date / Start Time, End Date / End Time
    # same logic as your original

    # 10a) Start Date
    currentElementSelector = (
        '.sapClientMDiv.sapClientControlsCoreControlProxy.sapClientMFormPane.saplsui-widthauto'
        ' > div > div > div > div:nth-of-type(7) > div > div > div'
        ' > div:nth-of-type(1)'
    )
    elementDescription = 'date'
    click_element(
        currentElementSelector, 
        elementDescription,
        maxAttempts=5,
        stopAfterClickBeforeCheckClick=SHORT_WAIT,
        stopAfterClickedSuccessfully=SHORT_WAIT
    )


    click_element(currentElementSelector, "Start Date")
    sd_cell = sheet1.cell(row_index, columnStartDateTime)
    if sd_cell.value:
        text_val = str(sd_cell.value).strip()
        el = driver.find_element(By.CSS_SELECTOR, currentElementSelector)
        action.move_to_element(el).pause(SHORT_WAIT).double_click().click() \
              .send_keys_to_element(el, Keys.BACKSPACE).pause(SHORT_WAIT) \
              .send_keys_to_element(el, text_val).pause(SHORT_WAIT) \
              .send_keys_to_element(el, Keys.TAB).pause(SHORT_WAIT).perform()
    time.sleep(SHORT_WAIT)

    # 10b) Start Time
    start_time_sel = (
        '.sapClientMDiv.sapClientControlsCoreControlProxy.sapClientMFormPane.saplsui-widthauto'
        ' > div > div > div > div:nth-of-type(7) > div > div > div'
        ' > div:nth-of-type(2)'
    )
    click_element(start_time_sel, "Start Time")
    el = driver.find_element(By.CSS_SELECTOR, start_time_sel)
    action.move_to_element(el).pause(SHORT_WAIT).double_click().click() \
          .send_keys_to_element(el, Keys.BACKSPACE).pause(SHORT_WAIT) \
          .send_keys_to_element(el, "8:00 am").pause(SHORT_WAIT) \
          .send_keys_to_element(el, Keys.TAB).pause(SHORT_WAIT).perform()
    time.sleep(SHORT_WAIT)

    # 10c) End Date
    end_date_sel = (
        '.sapClientMDiv.sapClientControlsCoreControlProxy.sapClientMFormPane.saplsui-widthauto'
        ' > div > div > div > div:nth-of-type(8) > div > div > div'
        ' > div:nth-of-type(1)'
    )
    click_element(end_date_sel, "End Date")
    
    sd_cell = sheet1.cell(row_index, columnStartDateTime)
    if sd_cell.value:
        text_val = str(sd_cell.value).strip()
        el = driver.find_element(By.CSS_SELECTOR, end_date_sel)
        action.move_to_element(el).pause(SHORT_WAIT).double_click().click() \
              .send_keys_to_element(el, Keys.BACKSPACE).pause(SHORT_WAIT) \
              .send_keys_to_element(el, text_val).pause(SHORT_WAIT) \
              .send_keys_to_element(el, Keys.TAB).pause(SHORT_WAIT).perform()
    time.sleep(SHORT_WAIT)

    # 10d) End Time
    end_time_sel = (
        '.sapClientMDiv.sapClientControlsCoreControlProxy.sapClientMFormPane.saplsui-widthauto'
        ' > div > div > div > div:nth-of-type(8) > div > div > div'
        ' > div:nth-of-type(2)'
    )
    click_element(end_time_sel, "End Time")    
    
    el = driver.find_element(By.CSS_SELECTOR, end_time_sel)
    action.move_to_element(el).pause(SHORT_WAIT).double_click().click() \
          .send_keys_to_element(el, Keys.BACKSPACE).pause(SHORT_WAIT) \
          .send_keys_to_element(el, "9:00 am").pause(SHORT_WAIT) \
          .send_keys_to_element(el, Keys.TAB).pause(SHORT_WAIT).perform()
    time.sleep(SHORT_WAIT)

    # 11) Internal Notes (if any)
    notes_val = sheet1.cell(row_index, columnNotes).value
    if notes_val:
        notes_sel = (
            '.sapUiRespGrid.sapUiRespGridHSpace0.sapUiRespGridVSpace0.sapUiFormResGridCont.sapUiRespGridOverflowHidden.sapUiRespGridMedia-Std-Desktop'
            ' > div:nth-of-type(11) > div > div [title="Editable area. Press F10 for toolbar."]'
            )
        currentElementSelector = (
            'textarea.sapMInputBaseInner.sapMTextAreaInner.k-content.k-raw-content'
            )
        currentElementSelector = 'td.k-editable-area'
        elementDescription = 'Internal Notes'
        click_element(
            currentElementSelector, 
            elementDescription,
            maxAttempts=5,
            stopAfterClickBeforeCheckClick=SHORT_WAIT,
            stopAfterClickedSuccessfully=SHORT_WAIT
            )
        time.sleep(SHORT_WAIT)
        action.send_keys(notes_val).perform()
        time.sleep(MED_WAIT)

    # 12) Save
    save_sel = (
        '.sapMPageFooter button.sapMBtnBase.sapMBtn.sapClientMQuickCreateAction .sapMBtnContent'
    )
    click_element(save_sel, "Save", 5)
    
    time.sleep(MED_WAIT*2)
    # 12a) Confirm we're back on the account page
    new_visit_sel = (
        '.sapUiVltCell.sapuiVltCell.sapClientMALPToolbarParent '
        '.sapMBtnBase.sapMBtn.width-button-form.sapMBarChild '
        '.sapMBtnContent'
    )    
    if check_for_element(new_visit_sel, "Confirm Save", 3):
        # 12b) Close account window
        close_sel = (
            '[data-sap-ui="mainShell-container"] '
            '.sapClientMMultiTabLayoutTopBar '
            '.sapClientMTabButton.sapClientMTabButtonSelected.'
            'sapMBarChild.sapClientMTabButtonIconVisible [title="Close"]'
        )
        click_element(close_sel, "Close Account Window", 5)
        time.sleep(MED_WAIT)
        click_element(close_sel, "Close Account Window", 5)
        time.sleep(MED_WAIT)

        # 12c) Go Home
        home_sel = '#__div2'
        click_element(home_sel, "Home", 3)
        home_check_sel = (
            '.sapUiUfdShell.sapUiUfdShellAnim.sapUiUfdShellHeadVisible.'
            'sapUiUfdShellCurtainHidden.sapUiUfdShellCurtainClosed.sapClientMShell.'
            'sapClientMShellLeftOpen.sapClientMShellDisableLetterBox.sapClientMShellMultiTabLayout.'
            'sapClientMShellWithHomePage #mainShell-cntnt'
        )
        if check_for_element(home_check_sel, "Home Page Check", 3):
            print(f"Activity in row {row_index} logged to SAP.\n")
            sheet1.cell(row=row_index, column=columnLoggedSuccessfully).value = "Activity Logged"
            report.save(reportPath)

def main_loop():
    # Open chrome://settings, set zoom, etc.
    driver.get('chrome://settings/')
    time.sleep(SHORT_WAIT)
    driver.execute_script('chrome.settingsPrivate.setDefaultZoom(1);')
    time.sleep(SHORT_WAIT)

    # Navigate to SAP Cloud
    driver.get("https://my363805-sso.crm.ondemand.com/sap/ap/ui/repository/SAP_UI/HTML5/newclient.html?sap-language=EN&sap-ui-language=en_us#")
    time.sleep(MED_WAIT)

    # Adjust window positions
    driver.set_window_position(-1000, 0)
    time.sleep(SHORT_WAIT)
    driver.maximize_window()
    time.sleep(SHORT_WAIT)

    global rowCurrent, rowMaxCount
    # Count rows
    row_count = 0
    for row in sheet1.iter_rows(values_only=True):
        if not all(cell is None for cell in row):
            row_count += 1
    rowMaxCount = row_count
    rowCurrent = 2  # Start from row 2 (header presumably on row 1)
    
    while rowCurrent <= rowMaxCount:
        print(f"\n=== Processing row {rowCurrent}/{rowMaxCount} ===\n")
        try:
            process_one_row(rowCurrent)
        except Exception as e:
            msg = f"Row {rowCurrent} encountered an error: {str(e)}"
            print(msg)
            logger.error(msg)
            sheet1.cell(rowCurrent, columnLoggedSuccessfully).value = f"ERROR: {str(e)}"
            report.save(reportPath)
        finally:
            rowCurrent += 1
            time.sleep(SHORT_WAIT)

    print("All rows processed.\n")
    report.close()

###############################################################################
# ENTRY POINT
###############################################################################

main_loop()
'''
try:
    main_loop()
except KeyboardInterrupt:
    print("\nKeyboardInterrupt detected. Saving partial progress and exiting...\n")
    logger.warning("KeyboardInterrupt detected. Saving partial progress.")
finally:
    # Let system sleep
    ctypes.windll.kernel32.SetThreadExecutionState(0x80000000)
    print("Closing browser...")
    driver.quit()
    print("All activity has been loaded into SAP. Script done.")
'''
